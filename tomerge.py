# -*- coding: UTF-8 -*-
import os
import openpyxl


def savexls(src, dst):
    srcwb = openpyxl.load_workbook(src)
    srcws = srcwb.get_sheet_by_name(srcwb.get_sheet_names()[0])
    if os.path.exists(dst):
        dstwb = openpyxl.load_workbook(dst)
        dstws = dstwb.get_sheet_by_name(dstwb.get_sheet_names()[0])
        dst_max_row = dstws.max_row  # 最大行数
    else:
        dst_max_row = 0
        openpyxl.Workbook().save(dst)
        dstwb = openpyxl.load_workbook(dst)
        dstws = dstwb.get_sheet_by_name(dstwb.get_sheet_names()[0])
    for i, row in enumerate(srcws.iter_rows()):
        for j, col in enumerate(row):
            dstws.cell(row=i + 1 + dst_max_row, column=j + 1, value=col.value)
    dstwb.save(dst)


def file_name(file_dir):
    L = []
    for root, dirs, files in os.walk(file_dir):
        for file in files:
            if os.path.splitext(file)[1] == '.xlsx':
                L.append(os.path.join(root, file))
    return L


def main():
    srcfile=file_name(os.getcwd())
    print(u'merge')
    dstfile = os.path.dirname(os.path.abspath(__file__)).split('\\')[-1] + u'.xlsx'
    assert not os.path.exists(dstfile), "file already existed"
    for i, x in enumerate(srcfile):
        savexls(x, dstfile)  # 尾插
    print(u'all done')


if __name__ == '__main__':
    main()
